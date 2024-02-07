import json
import openpyxl
import logging

# Configuration
JSON_DB_PATH = './db.json'
EXCEL_DB_PATH = './database.xlsx'
NEW_JSON_DB_PATH = './db_new.json'

# Initialize logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

def load_json_data(file_path):
    try:
        with open(file_path, 'r') as file:
            data = json.load(file)
        logging.info(f"Loaded JSON data with {len(data)} records.")
        return data
    except Exception as e:
        logging.error(f"Failed to load JSON data: {e}")
        return None

def load_excel_data(file_path):
    try:
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active
        data = {}
        for i in range(2, sheet.max_row + 1):
            episode_num = sheet.cell(row=i, column=8).value
            if episode_num is not None:
                episode_num = str(int(episode_num)).strip()  # Convert to integer and then to string
                data[episode_num] = (sheet.cell(row=i, column=14).value, sheet.cell(row=i, column=15).value)
        logging.info(f"Loaded Excel data with {len(data)} records.")
        return data
    except Exception as e:
        logging.error(f"Failed to load Excel data: {e}")
        return None

def update_json_with_excel_data(json_data, excel_data):
    updated_count = 0
    for episode in json_data:
        episode_num = episode.get('episode_number', '').strip()
        if episode_num in excel_data:
            episode['human_post'], episode['gpt_summary'] = excel_data[episode_num]
            updated_count += 1
    logging.info(f"Updated {updated_count} records in JSON data.")

def write_json_data(file_path, data):
    try:
        with open(file_path, 'w') as file:
            json.dump(data, file, indent=4)
        logging.info(f"Written updated JSON data to {file_path}")
    except Exception as e:
        logging.error(f"Failed to write JSON data: {e}")

def main():
    json_data = load_json_data(JSON_DB_PATH)
    excel_data = load_excel_data(EXCEL_DB_PATH)
    if json_data is None or excel_data is None:
        return

    update_json_with_excel_data(json_data, excel_data)
    write_json_data(NEW_JSON_DB_PATH, json_data)

if __name__ == "__main__":
    main()
