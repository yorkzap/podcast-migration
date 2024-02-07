import openpyxl
import requests
import time
import logging
import json

# Configurations
MAX_POSTS = 161
MAX_INPUT_TOKENS = 8000  # This will be used to truncate the raw data
SLEEP_DURATION = 2.1  # in seconds
OPENAI_API_KEY = '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'  # Replace with your actual API key
FILE_PATH = './database.xlsx'
NEW_FILE_PATH = './database2.xlsx'

# Initialize logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

def read_excel_columns(file_path, column_letters):
    try:
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active
        return {column: [(cell.row, cell.value) for cell in sheet[column] if cell.value is not None] for column in column_letters}
    except Exception as e:
        logging.error(f"Failed to open {file_path}: {e}")
        return None

def write_to_excel(file_path, column_letter, row_number, data):
    try:
        # Load the existing workbook or create a new one if it doesn't exist
        try:
            workbook = openpyxl.load_workbook(file_path)
        except FileNotFoundError:
            workbook = openpyxl.Workbook()

        sheet = workbook.active
        sheet[f'{column_letter}{row_number}'] = data
        workbook.save(file_path)
    except Exception as e:
        logging.error(f"Failed to write to {file_path}: {e}")

def truncate_text(text, max_tokens):
    return ' '.join(text.split()[:max_tokens])

def generate_summary(raw_data):
    truncated_data = truncate_text(raw_data, MAX_INPUT_TOKENS)
    system_message = (
        f"Craft a concise gem of 250 characters max"
        "Tone: motivational and empathetic and like the information given in the context given at the end"
        "Style: casual, informative and persuasive. "
        "Voice: confident, experienced. "
        "Write in a direct manner with varying sentence lengths and conversational tone. "
        "Avoid jargon and be clear and simple. "
        "Remember this is the juice of the entire data we have given below:"
        "Below is all the context: " + truncated_data
    )
    
    data = {
        "model": "gpt-3.5-turbo-1106",
        "messages": [
            {"role": "system", "content": system_message}
        ]
    }
    headers = {
        "Content-Type": "application/json",
        "Authorization": f"Bearer {OPENAI_API_KEY}"
    }

    try:
        response = requests.post("https://api.openai.com/v1/chat/completions", headers=headers, data=json.dumps(data))
        response.raise_for_status()
        return response.json()['choices'][0]['message']['content']
    except requests.exceptions.RequestException as e:
        logging.error(f"Error during API call: {e}")
        return None

def get_preferred_column_data(columns_data):
    preferred_data = columns_data.get('N') or columns_data.get('L') or columns_data.get('B')
    return preferred_data if preferred_data else []

def main():
    logging.info("Starting the summary generation script...")

    start_row = input("Enter the starting row number: ")
    try:
        start_row = int(start_row)
    except ValueError:
        logging.error("Invalid input. Please enter a valid row number.")
        return

    columns_data = read_excel_columns(FILE_PATH, ['N', 'L', 'B'])
    if columns_data is None:
        return

    preferred_column_data = get_preferred_column_data(columns_data)
    post_count = 0

    for row_number, raw_data in preferred_column_data:
        if row_number < start_row or not raw_data:
            continue

        if post_count >= MAX_POSTS:
            logging.info("Reached maximum number of summaries to generate.")
            break

        logging.info(f"Row {row_number}: Generating summary...")
        summary = generate_summary(raw_data)
        if summary:
            write_to_excel(NEW_FILE_PATH, 'O', row_number, summary)
            post_count += 1
            logging.info(f"Row {row_number}: Summary written successfully. | {summary}")
        else:
            logging.warning(f"Row {row_number}: No summary generated.")

        time.sleep(SLEEP_DURATION)

    logging.info("Script completed.")

if __name__ == "__main__":
    main()
